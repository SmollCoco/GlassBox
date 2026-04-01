"""Excel I/O utilities."""

from __future__ import annotations

from typing import Any

import numpy as np

from ..core.dataframe import DataFrame
from ..utils.dtypes import infer_dtype, parse_scalar


def read_excel(filepath: str, sheet_name: int | str = 0) -> DataFrame:
    """Read an Excel worksheet into a DataFrame.

    Parameters
    ----------
    filepath : str
        Path to Excel file.
    sheet_name : int or str, default 0
        Sheet index or name.

    Returns
    -------
    DataFrame
        Parsed DataFrame.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(filepath, data_only=True)
    if isinstance(sheet_name, int):
        sheet = workbook.worksheets[sheet_name]
    else:
        sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return DataFrame({})
    columns = [str(col) for col in rows[0]]
    data_rows = rows[1:]
    col_values = {name: [] for name in columns}
    for row in data_rows:
        if len(row) != len(columns):
            raise ValueError("Row length does not match header.")
        for name, value in zip(columns, row):
            col_values[name].append(parse_scalar(value))
    data = {}
    for name, values in col_values.items():
        dtype = infer_dtype(values)
        data[name] = np.array(values, dtype=dtype)
    return DataFrame(data, columns=columns)


def write_excel(frame: DataFrame, filepath: str, sheet_name: str = "Sheet1") -> None:
    """Write a DataFrame to an Excel worksheet.

    Parameters
    ----------
    frame : DataFrame
        DataFrame to write.
    filepath : str
        Output path.
    sheet_name : str, default 'Sheet1'
        Worksheet name.
    """
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(frame.columns)
    for row_idx in range(len(frame.index)):
        row = [frame._data[col][row_idx] for col in frame.columns]
        sheet.append(row)
    workbook.save(filepath)
